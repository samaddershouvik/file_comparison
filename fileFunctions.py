import os
import difflib
import pdfplumber
from docx import Document
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import shutil
import re
from openpyxl.styles import Alignment, Border, Side

productName = "Bajaj Allianz Life Ins. Co. Ltd. "


def extract_text_from_file(file_path):
    """Extract text from a file along with page numbers and titles."""
    page_data = []  # Store (page_number, page_title, text) tuples

    if file_path.endswith(".txt"):
        with open(file_path, "r") as file:
            text = file.readlines()
            title = next(
                (line.strip() for line in text if line.strip().isupper()), "Unknown"
            )
            page_data.append((1, title, text))

    elif file_path.endswith(".pdf"):
        with pdfplumber.open(file_path) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                lines = page.extract_text().split("\n") if page.extract_text() else []
                title = next(
                    (line.strip() for line in lines if line.strip().isupper()),
                    "Unknown",
                )
                page_data.append((page_number, title, lines))

    elif file_path.endswith(".docx"):
        doc = Document(file_path)
        text = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        page_number = 1  # No real pagination in DOCX
        title = next((line for line in text if line.isupper()), "Unknown")
        page_data.append((page_number, title, text))

    else:
        raise ValueError(f"Unsupported file format: {file_path}")

    return page_data  # List of (page_number, page_title, text)


def extract_date_from_filename(filename):
    """Extract YYYYMMDD date pattern from the filename if present."""
    match = re.search(r"\b(20\d{6})\b", filename)  # Matches 2024XXXX format
    return match.group(1) if match else ""  # Return the extracted date or empty string


def format_excel_sheet(worksheet):
    """Format the Excel sheet by merging Product Name, Product UIN, and Sample Affected columns
    and applying bordered styling for each change row."""

    # Define border styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    start_row = 2  # Start from second row (first row is headers)
    end_row = worksheet.max_row  # Last row with data

    # Merge "Product Name", "Product UIN", and "Sample Affected" vertically
    worksheet.merge_cells(
        start_row=start_row, end_row=end_row, start_column=1, end_column=1
    )  # Product Name
    worksheet.merge_cells(
        start_row=start_row, end_row=end_row, start_column=2, end_column=2
    )  # Product UIN
    worksheet.merge_cells(
        start_row=start_row, end_row=end_row, start_column=3, end_column=3
    )  # Sample Affected

    # Apply vertical alignment to center for merged columns using a loop
    for col in [1, 2, 3]:  # Columns: Product Name, Product UIN, Sample Affected
        worksheet.cell(row=start_row, column=col).alignment = Alignment(
            vertical="center"
        )

    # Apply thick border to merged columns
    for row in range(start_row, end_row + 1):
        for col in [1, 2, 3]:  # Columns: Product Name, Product UIN, Sample Affected
            worksheet.cell(row=row, column=col).border = thin_border

    # Apply borders for individual change rows (for content)
    for row in range(start_row, end_row + 1):
        for col in range(
            4, worksheet.max_column + 1
        ):  # Skip first three merged columns
            worksheet.cell(row=row, column=col).border = (
                thin_border  # Normal row border
            )


def generate_excel_report(mother_file, comparison_file, output_dir):
    """Generate an Excel report with a custom table format."""
    try:
        mother_data = extract_text_from_file(mother_file)
        comparison_data = extract_text_from_file(comparison_file)

        report_data = []
        mother_pages = {
            page_num: (title, text) for page_num, title, text in mother_data
        }
        comparison_pages = {
            page_num: (title, text) for page_num, title, text in comparison_data
        }

        all_page_nums = sorted(set(mother_pages.keys()) | set(comparison_pages.keys()))

        comparison_filename = os.path.basename(comparison_file)
        extracted_date = extract_date_from_filename(comparison_filename)

        for page_num in all_page_nums:
            mother_title, mother_content = mother_pages.get(page_num, ("Unknown", []))
            comparison_title, comparison_content = comparison_pages.get(
                page_num, ("Unknown", [])
            )

            diff = difflib.ndiff(mother_content, comparison_content)

            for line in diff:
                if line.startswith("- "):  # Removed from mother file
                    report_data.append(
                        {
                            "Product Name": productName,  # Static value
                            "Product UIN": "",  # Leave blank
                            "Sample Affected": extracted_date,  # Extracted Date
                            "Observation Category": "Available in Filed Copy but missing in Customer Copy",  # Status
                            "Part": mother_title,  # Page Title
                            "Page Number": page_num,
                            f"Filed Copy ({os.path.basename(mother_file)}) Content": line[2:].strip(),
                            f"Customer Copy ({os.path.basename(comparison_file)}) Content": "",
                        }
                    )
                elif line.startswith("+ "):  # Added in comparison file
                    report_data.append(
                        {
                            "Product Name": productName,  # Static value
                            "Product UIN": "",  # Leave blank
                            "Sample Affected": extracted_date,  # Extracted Date
                            "Observation Category": "Mismatch of content between Filed Copy and customer copy",  # Status
                            "Part": comparison_title,  # Page Title
                            "Page Number": page_num,
                            f"Filed Copy ({os.path.basename(mother_file)}) Content": "",
                            f"Customer Copy ({os.path.basename(comparison_file)}) Content": line[2:].strip(),
                        }
                    )

        if not report_data:
            print(f"No differences found between {mother_file} and {comparison_file}.")
            return

        df = pd.DataFrame(report_data)

        output_file = os.path.join(output_dir, f"{comparison_filename} Report.xlsx")

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Differences")

            worksheet = writer.sheets["Differences"]

            format_excel_sheet(worksheet)

            red_fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            green_fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )

            for row in worksheet.iter_rows(
                min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)
            ):
                if (
                    row[3].value
                    == "Available in Filed Copy but missing in Customer Copy"
                ):  # Observation Category column index
                    for cell in row:
                        cell.fill = red_fill
                elif (
                    row[3].value
                    == "Mismatch of content between Filed Copy and customer copy"
                ):
                    for cell in row:
                        cell.fill = green_fill

            for col_idx, col_cells in enumerate(worksheet.columns, start=1):
                max_length = (
                    max(len(str(cell.value)) for cell in col_cells if cell.value) + 2
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = (
                    max_length
                )

        print(f"Differences report saved to {output_file}")

    except Exception as e:
        print(f"An error occurred: {e}")


def compare_with_mother_file(mother_file_path, comparison_folder):
    """
    Compare the mother file with all files in a specified folder and move processed files.
    """
    # Validate the comparison folder
    if not os.path.isdir(comparison_folder):
        print("The second input must be a valid folder containing files to compare.")
        return

    # Create an output directory for the reports
    parent_dir = os.path.dirname(mother_file_path)
    output_dir = os.path.join(parent_dir, "Comparison_Reports")
    os.makedirs(output_dir, exist_ok=True)

    # Process and compare files
    for file_name in os.listdir(comparison_folder):
        comparison_file_path = os.path.join(comparison_folder, file_name)
        if os.path.isfile(comparison_file_path):  # Ensure it's a file
            generate_excel_report(mother_file_path, comparison_file_path, output_dir)

    # Move processed files to the 'processed' folder
    processed_folder = os.path.join(comparison_folder, "processed")
    move_to_processed_folder(comparison_folder, processed_folder)


def move_to_processed_folder(comparison_folder, processed_folder):
    """
    Move all files from the comparison folder to the processed folder after comparison.
    """
    # Create the processed folder if it doesn't exist
    os.makedirs(processed_folder, exist_ok=True)

    # Iterate through files in the comparison folder
    for file_name in os.listdir(comparison_folder):
        file_path = os.path.join(comparison_folder, file_name)
        if os.path.isfile(file_path):  # Ensure it's a file
            # Move the file to the processed folder
            shutil.move(file_path, os.path.join(processed_folder, file_name))
            print(f"Moved file: {file_name} to {processed_folder}")
